# -*- coding: utf-8 -*-
import openpyxl
import re
from openpyxl.utils import get_column_letter

path_to_file = 'base.xlsx'


def check_url():
    search_text = input(str('Print url: '))
    search_text = search_text.lower()
    print('Search:', search_text)

    wb = openpyxl.load_workbook(path_to_file)  # Грузим наш прайс-лист
    sheets_list = wb.sheetnames  # Получаем список всех листов в файле
    sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
    row_max = sheet_active.max_row  # Получаем количество столбцов

    column_max = sheet_active.max_column  # Получаем количество строк

    print('В файле:', path_to_file, '\n Cтолбцов:', row_max, '\n Колонок:', column_max)

    row_min = 1  # Переменная, отвечающая за номер строки
    column_min = 1  # Переменная, отвечающая за номер столбца

    row_min_min = row_min
    row_max_max = row_max
    match = None
    while row_min_min <= row_max_max:
        row_min_min = str(row_min_min)

        word_column = get_column_letter(column_min)
        word_column = str(word_column)
        word_cell = word_column + row_min_min

        data_from_cell = sheet_active[word_cell].value
        data_from_cell = str(data_from_cell)
        regular = search_text
        result = re.match(regular, data_from_cell)
        if result is not None:
            match = +1
        row_min_min = int(row_min_min)
        row_min_min = row_min_min + 1

    column_min = column_min + 1
    if match is not None:
        print('Yes')
    else:
        print('None')


if __name__ == '__main__':
    check_url()
